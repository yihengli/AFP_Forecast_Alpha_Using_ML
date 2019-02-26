import logging
import os

import yaml
from tqdm import tqdm, tqdm_notebook

import colorlog
import colorama
from skopt.space import Integer, Categorical, Real


def load_search_cv_config(path_config):
    """
    Load the configurations recorded in the config file.

    Numpy related functions will be translated and executed.

    Parameters
    ----------
    path_config : str
        The abs path to the `yaml` config file

    Returns
    -------
    dict
        The dict of kwargs that can be directly used by SearchCV objects.
    """
    if path_config is None:
        return path_config

    with open(path_config, 'r') as config:
        settings = yaml.load(config)
        for k, v in settings['search_spaces'].items():
            if isinstance(v, dict) and 'Real' in v.keys():
                if 'log-uniform' in v:
                    settings['search_spaces'][k] = Real(v['Real'][0],
                                                        v['Real'][1],
                                                        'log-uniform')
                else:
                    settings['search_spaces'][k] = Real(v['Real'][0],
                                                        v['Real'][1])
            elif isinstance(v, dict) and 'Integer' in v.keys():
                settings['search_spaces'][k] = Integer(v['Integer'][0],
                                                       v['Integer'][1])
            elif isinstance(v, dict) and 'Categorical' in v.keys():
                settings['search_spaces'][k] = Categorical(v['Categorical'])
        return settings


def in_ipynb():
    """
    Check if the current environment is IPython Notebook

    Note, Spyder terminal is also using ZMQShell but cannot render Widget.

    Returns
    -------
    bool
        True if the current env is Jupyter notebook
    """
    try:
        zmq_status = str(type(get_ipython())) == "<class 'ipykernel.zmqshell.ZMQInteractiveShell'>"  # noqa E501
        spyder_status = any('SPYDER' in name for name in os.environ)
        return zmq_status and not spyder_status

    except NameError:
        return False


def get_tqdm():
    """
    Get proper tqdm progress bar instance based on if the current env is
    Jupyter notebook

    Note, Windows system doesn't supprot default progress bar effects

    Returns
    -------
    type, bool
        either tqdm or tqdm_notebook, the second arg will be ascii option
    """
    ascii = True if os.name == 'nt' else False

    if in_ipynb():
        # Jupyter notebook can always handle ascii correctly
        return tqdm_notebook, False
    else:
        return tqdm, ascii


class TqdmHandler(logging.StreamHandler):
    """
    A logging output handler that will use TQDM write method to live with
    tqdm progress bar
    """

    def __init__(self):
        colorlog.StreamHandler.__init__(self)

    def emit(self, record):
        msg = self.format(record)
        tqdm.write(msg)


def get_logger(name='afp'):
    """
    Get a proper logger object for this project. The logger has been formatted
    with colors and should be able to use together with TQDM progressbars

    Parameters
    ----------
    name : str, optional
        The name of the logger (the default is 'afp')

    Returns
    -------
    logging.RootLogger
        The logger object for this project
    """

    logger = colorlog.getLogger(name)

    # Disable colorama if in jupyter notebook env
    if in_ipynb():
        colorama.deinit()

    if logger.hasHandlers():
        logger.handlers.clear()

    handler = TqdmHandler()
    handler.setFormatter(
        colorlog.ColoredFormatter(
            '%(log_color)s%(asctime)s - %(log_color)s %(levelname)-6s%(reset)s %(white)s%(message)s'))  # noqa E501
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    return logger


def _get_between(df, start, end):
    return df[(df.index >= start) & (df.index <= end)]


def _get_common_nonnull(col1, col2):
    idx_common = col1.dropna().index.intersection(col2.dropna().index)
    return col1.loc[idx_common], col2.loc[idx_common]


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to DataFrame.to_excel()
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
